#!/usr/bin/env python3
"""
Universal News Scraper for Inquirer, Business Mirror, and Philstar
Scrapes all sources, saves Excel files, and uploads to Azure Blob Storage
Optimized for GitHub Actions workflow with enhanced error handling
"""

import os
import sys
import pandas as pd
from datetime import datetime
import importlib
from dotenv import load_dotenv

# Load environment variables from .env file (for local development)
load_dotenv()

def import_scraper(module_name, function_name):
    """Import scraper functions with error handling"""
    try:
        module = importlib.import_module(module_name)
        return getattr(module, function_name)
    except Exception as e:
        print(f"❌ Error importing {module_name}.{function_name}: {e}")
        return None

def validate_azure_environment():
    """Validate Azure environment configuration for both local and CI environments"""
    azure_conn = os.getenv('AZURE_CONNECTION_STRING')
    azure_container = os.getenv('AZURE_CONTAINER_NAME')
    
    if not azure_conn:
        print("❌ AZURE_CONNECTION_STRING environment variable not found")
        print("💡 For local development: Check your .env file")
        print("💡 For GitHub Actions: Check repository secrets")
        return False
    
    if not azure_container:
        print("❌ AZURE_CONTAINER_NAME environment variable not found")
        print("💡 For local development: Check your .env file")
        print("💡 For GitHub Actions: Check repository secrets")
        return False
    
    print("✅ Azure Blob Storage configuration validated")
    print(f"📦 Container: {azure_container}")
    return True

def main():
    """Main orchestrator function with enhanced error handling for GitHub Actions"""
    print("🤖 Universal News Scraper - GitHub Actions Optimized")
    print(f"⏰ Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Validate Azure environment first
    if not validate_azure_environment():
        print("❌ Azure environment validation failed - cannot proceed")
        sys.exit(1)
    
    # Import enhanced scrapers with dynamic date filtering and improved extraction
    print("\n📰 Importing news scraper modules...")
    
    # Import all three scrapers
    inquirer_scraper = import_scraper('scrape_inquirer', 'scrape_inquirer_news')
    inquirer_upload = import_scraper('scrape_inquirer', 'upload_to_azure_blob')
    
    businessmirror_scraper = import_scraper('scrape_businessmirror_fixed', 'scrape_businessmirror_news')
    businessmirror_upload = import_scraper('scrape_businessmirror_fixed', 'upload_to_azure_blob')
    
    philstar_scraper = import_scraper('scrape_philstar_improved', 'scrape_philstar_news')
    philstar_upload = import_scraper('scrape_philstar_improved', 'upload_to_azure_blob')
    
    # Check if all imports were successful
    missing_functions = []
    if not inquirer_scraper: missing_functions.append("inquirer_scraper")
    if not inquirer_upload: missing_functions.append("inquirer_upload")
    if not businessmirror_scraper: missing_functions.append("businessmirror_scraper")
    if not businessmirror_upload: missing_functions.append("businessmirror_upload")
    if not philstar_scraper: missing_functions.append("philstar_scraper")
    if not philstar_upload: missing_functions.append("philstar_upload")
    
    if missing_functions:
        print(f"❌ Could not import required functions: {', '.join(missing_functions)}")
        print("💡 Ensure all scraper files are present and contain the required functions")
        sys.exit(1)
    
    print("✅ All scraper modules imported successfully")
    print("🚀 Enhanced Features:")
    print("   • Dynamic date filtering (today & yesterday)")
    print("   • Enhanced anti-bot measures with modern user agents")
    print("   • Retry logic for Azure uploads")
    print("   • Proper error handling for GitHub Actions")

    # Track success/failure for exit code
    scraping_errors = 0

    # Scrape Inquirer
    print("\n==============================")
    print("🔍 Scraping Inquirer Business News...")
    try:
        inquirer_news = inquirer_scraper()
        if inquirer_news:
            import pandas as pd
            df_inq = pd.DataFrame(inquirer_news)
            inq_filename = "inquirer_news.xlsx"
            df_inq.to_excel(inq_filename, index=False)
            # Add Excel table named 'NewsTable'
            try:
                from openpyxl import load_workbook
                from openpyxl.worksheet.table import Table, TableStyleInfo
                wb = load_workbook(inq_filename)
                ws = wb.active
                nrows = ws.max_row
                ncols = ws.max_column
                col_letters = ws.cell(row=1, column=ncols).column_letter
                table_ref = f"A1:{col_letters}{nrows}"
                table = Table(displayName="NewsTable", ref=table_ref)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                table.tableStyleInfo = style
                ws.add_table(table)
                wb.save(inq_filename)
                print(f"📊 Saved {len(df_inq)} Inquirer news items to {inq_filename} (with table 'NewsTable')")
            except Exception as e:
                print(f"⚠️ Could not add Excel table to Inquirer file: {e}")
                print(f"📊 Saved {len(df_inq)} Inquirer news items to {inq_filename}")
            
            print("☁️ Uploading Inquirer news to Azure...")
            upload_success = inquirer_upload(inq_filename, inq_filename)
            if not upload_success:
                print("❌ Failed to upload Inquirer news to Azure")
                scraping_errors += 1
            else:
                print("✅ Inquirer news uploaded successfully")
        else:
            print("❌ No Inquirer news found.")
            scraping_errors += 1
    except Exception as e:
        print(f"❌ Inquirer scraping failed: {str(e)}")
        scraping_errors += 1

    # Scrape Business Mirror
    print("\n==============================")
    print("🔍 Scraping Business Mirror News...")
    try:
        # Check if file is locked before proceeding
        bm_filename = "businessmirror_news.xlsx"
        if os.path.exists(bm_filename):
            try:
                # Test if we can access the file
                with open(bm_filename, 'r+b') as test_file:
                    pass
            except PermissionError:
                print(f"⚠️  File {bm_filename} is currently locked. Please close any Excel applications and try again.")
                print("   Skipping Business Mirror scraping to continue with other sources...")
                bm_news = None
            except:
                print(f"   Removing existing file: {bm_filename}")
                os.remove(bm_filename)
        
        if 'bm_news' not in locals() or bm_news is None:
            bm_news = businessmirror_scraper()
        
        if bm_news:
            import pandas as pd
            df_bm = pd.DataFrame(bm_news)
            df_bm.to_excel(bm_filename, index=False)
            # Add Excel table named 'NewsTable1'
            try:
                from openpyxl import load_workbook
                from openpyxl.worksheet.table import Table, TableStyleInfo
                wb = load_workbook(bm_filename)
                ws = wb.active
                nrows = ws.max_row
                ncols = ws.max_column
                col_letters = ws.cell(row=1, column=ncols).column_letter
                table_ref = f"A1:{col_letters}{nrows}"
                table = Table(displayName="NewsTable1", ref=table_ref)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                table.tableStyleInfo = style
                ws.add_table(table)
                wb.save(bm_filename)
                print(f"📊 Saved {len(df_bm)} Business Mirror news items to {bm_filename} (with table 'NewsTable1')")
            except Exception as e:
                print(f"⚠️ Could not add Excel table to Business Mirror file: {e}")
                print(f"📊 Saved {len(df_bm)} Business Mirror news items to {bm_filename}")
            
            print("☁️ Uploading Business Mirror news to Azure...")
            upload_success = businessmirror_upload(bm_filename, bm_filename)
            if not upload_success:
                print("❌ Failed to upload Business Mirror news to Azure")
                scraping_errors += 1
            else:
                print("✅ Business Mirror news uploaded successfully")
        else:
            print("❌ No Business Mirror news found.")
            scraping_errors += 1
    except Exception as e:
        print(f"❌ Business Mirror scraping failed: {str(e)}")
        if "Permission denied" in str(e):
            print("   💡 Suggestion: Close any Excel files and ensure no applications are using the output file")
        scraping_errors += 1

    # Scrape Philstar
    print("\n==============================")
    print("🔍 Scraping Philstar Business News...")
    try:
        philstar_news = philstar_scraper()
        if philstar_news:
            import pandas as pd
            df_philstar = pd.DataFrame(philstar_news)
            philstar_filename = "philstar_news.xlsx"
            df_philstar.to_excel(philstar_filename, index=False)
            # Add Excel table named 'NewsTable2'
            try:
                from openpyxl import load_workbook
                from openpyxl.worksheet.table import Table, TableStyleInfo
                wb = load_workbook(philstar_filename)
                ws = wb.active
                nrows = ws.max_row
                ncols = ws.max_column
                col_letters = ws.cell(row=1, column=ncols).column_letter
                table_ref = f"A1:{col_letters}{nrows}"
                table = Table(displayName="NewsTable2", ref=table_ref)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                table.tableStyleInfo = style
                ws.add_table(table)
                wb.save(philstar_filename)
                print(f"📊 Saved {len(df_philstar)} Philstar news items to {philstar_filename} (with table 'NewsTable2')")
            except Exception as e:
                print(f"⚠️ Could not add Excel table to Philstar file: {e}")
                print(f"📊 Saved {len(df_philstar)} Philstar news items to {philstar_filename}")
            
            print("☁️ Uploading Philstar news to Azure...")
            upload_success = philstar_upload(philstar_filename, philstar_filename)
            if not upload_success:
                print("❌ Failed to upload Philstar news to Azure")
                scraping_errors += 1
            else:
                print("✅ Philstar news uploaded successfully")
        else:
            print("❌ No Philstar news found.")
            scraping_errors += 1
    except Exception as e:
        print(f"❌ Philstar scraping failed: {str(e)}")
        scraping_errors += 1

    # Final status and exit code
    print("\n==============================")
    if scraping_errors == 0:
        print("✅ Enhanced Universal News Scraping Complete!")
        print("🎯 All improvements implemented:")
        print("   ✓ Dynamic date filtering (adapts to any date)")
        print("   ✓ Business Mirror: More articles + proper categories")
        print("   ✓ Inquirer: Better date filtering accuracy")
        print("   ✓ Philstar: Infinite scroll simulation")
        print("   ✓ Enhanced error handling for GitHub Actions")
        print("   ✓ Retry logic for Azure uploads")
        print(f"⏰ Finished at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    else:
        print(f"⚠️ Universal News Scraping completed with {scraping_errors} error(s)")
        print("💡 Check the logs above for specific error details")
        print(f"⏰ Finished at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        sys.exit(1)  # Exit with error code for GitHub Actions

if __name__ == "__main__":
    main()
