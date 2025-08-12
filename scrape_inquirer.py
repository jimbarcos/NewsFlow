import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime, timedelta
import time
import re
from textblob import TextBlob
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime, timedelta
import time
import re
import os
from dotenv import load_dotenv
from azure.storage.blob import BlobServiceClient
try:
    from textblob import TextBlob
    from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
    SENTIMENT_AVAILABLE = True
except ImportError:
    SENTIMENT_AVAILABLE = False
    print("Note: Sentiment analysis libraries not available. Install with: pip install textblob vaderSentiment")

# Load environment variables
load_dotenv()

# Initialize sentiment analyzer
if SENTIMENT_AVAILABLE:
    analyzer = SentimentIntensityAnalyzer()

def get_sentiment_analysis(text):
    """Analyze sentiment of text using both TextBlob and VADER"""
    if not SENTIMENT_AVAILABLE or not text:
        return {
            'sentiment_score': 0.0,
            'sentiment_label': 'Neutral',
            'emotion': 'Neutral'
        }
    
    try:
        # TextBlob analysis
        blob = TextBlob(text)
        textblob_score = blob.sentiment.polarity
        
        # VADER analysis  
        vader_scores = analyzer.polarity_scores(text)
        vader_compound = vader_scores['compound']
        
        # Combine scores (average)
        combined_score = (textblob_score + vader_compound) / 2
        
        # Determine sentiment label
        if combined_score >= 0.1:
            sentiment_label = 'Positive'
            emotion = 'Optimistic' if combined_score > 0.3 else 'Positive'
        elif combined_score <= -0.1:
            sentiment_label = 'Negative'
            emotion = 'Concerning' if combined_score < -0.3 else 'Negative'
        else:
            sentiment_label = 'Neutral'
            emotion = 'Neutral'
        
        return {
            'sentiment_score': round(combined_score, 3),
            'sentiment_label': sentiment_label,
            'emotion': emotion
        }
    except Exception as e:
        print(f"Sentiment analysis error: {e}")
        return {
            'sentiment_score': 0.0,
            'sentiment_label': 'Neutral',
            'emotion': 'Neutral'
        }

def categorize_news(title, description=""):
    """Categorize news based on keywords in title and description"""
    text = (title + " " + description).lower()
    
    # Define category keywords
    categories = {
        'Banking & Finance': ['bank', 'financial', 'loan', 'credit', 'investment', 'fund', 'bsp', 'interest rate', 'monetary'],
        'Stock Market': ['stock', 'share', 'psei', 'market', 'trading', 'equity', 'index', 'surge', 'rally', 'decline'],
        'Energy & Utilities': ['fuel', 'oil', 'gas', 'energy', 'power', 'electricity', 'meralco', 'utility', 'renewable'],
        'Real Estate': ['property', 'real estate', 'housing', 'construction', 'development', 'land', 'residential'],
        'Technology': ['tech', 'digital', 'innovation', 'software', 'app', 'platform', 'online', 'cyber'],
        'Infrastructure': ['infrastructure', 'transport', 'road', 'bridge', 'airport', 'port', 'railway', 'construction'],
        'Retail & Consumer': ['retail', 'consumer', 'shopping', 'mall', 'store', 'sales', 'product', 'brand'],
        'Manufacturing': ['manufacturing', 'factory', 'production', 'industrial', 'export', 'import', 'goods'],
        'Agriculture': ['agriculture', 'farming', 'crop', 'rice', 'food', 'agricultural', 'fishery'],
        'Government & Policy': ['government', 'policy', 'regulation', 'law', 'sec', 'bir', 'dof', 'congress', 'senate'],
        'International Trade': ['trade', 'export', 'import', 'international', 'global', 'foreign', 'overseas'],
        'Economic Indicators': ['gdp', 'inflation', 'growth', 'economy', 'economic', 'recession', 'recovery']
    }
    
    # Check for category matches
    for category, keywords in categories.items():
        if any(keyword in text for keyword in keywords):
            return category
    
    return 'General Business'

def extract_description(link_element):
    """Extract description/snippet from the article's parent elements"""
    try:
        # Look for description in various parent elements
        parent = link_element.find_parent()
        description = ""
        
        # Try to find description in siblings or parent
        for _ in range(3):  # Check up to 3 parent levels
            if parent:
                # Look for common description classes/tags
                desc_selectors = [
                    '.excerpt', '.summary', '.description', '.content',
                    'p', '.entry-content', '.post-excerpt'
                ]
                
                for selector in desc_selectors:
                    desc_elem = parent.find(class_=selector) or parent.find(selector.replace('.', ''))
                    if desc_elem:
                        desc_text = desc_elem.get_text(strip=True)
                        if len(desc_text) > 20 and len(desc_text) < 300:
                            description = desc_text
                            break
                
                if description:
                    break
                parent = parent.find_parent()
        
        # Clean and limit description
        if description:
            description = re.sub(r'\s+', ' ', description)  # Remove extra whitespace
            description = description[:200] + "..." if len(description) > 200 else description
        
        return description
    except Exception:
        return ""

def extract_actual_article_date(url):
    """Extract actual publication date by visiting the article URL"""
    try:
        if not url or not url.startswith('http'):
            return None
            
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
        }
        
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code != 200:
            return None
            
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Method 1: Look for publication date in meta tags
        meta_selectors = [
            'meta[property="article:published_time"]',
            'meta[name="date"]',
            'meta[name="publish_date"]',
            'meta[property="og:published_time"]',
            'meta[name="publication_date"]'
        ]
        
        for selector in meta_selectors:
            meta_elem = soup.select_one(selector)
            if meta_elem:
                content = meta_elem.get('content')
                if content:
                    # Try to parse ISO date
                    try:
                        if 'T' in content or '+' in content:
                            # ISO format
                            parsed_date = datetime.fromisoformat(content.replace('Z', '+00:00'))
                            return parsed_date.strftime("%B %d, %Y")
                    except:
                        pass
        
        # Method 2: Look for date in article byline/header
        byline_patterns = [
            r'(\d{1,2}:\d{2}\s*(?:AM|PM))\s*(\w+)\s*(\d{1,2}),?\s*(\d{4})',  # "02:01 AM August 06, 2025"
            r'(\w+)\s*(\d{1,2}),?\s*(\d{4})\s*/\s*(\d{1,2}:\d{2}\s*(?:AM|PM))',  # "August 06, 2025 / 02:01 AM"
            r'(\w+)\s+(\d{1,2}),?\s+(\d{4})',  # "August 06, 2025"
            r'(\d{1,2})/(\d{1,2})/(\d{4})',  # "08/06/2025"
        ]
        
        # Get full page text for pattern matching
        page_text = soup.get_text()
        
        for pattern in byline_patterns:
            matches = re.findall(pattern, page_text, re.IGNORECASE)
            if matches:
                match = matches[0]
                
                # Handle different pattern formats
                if len(match) >= 3:
                    try:
                        # Pattern: "02:01 AM August 06, 2025"
                        if len(match) == 4 and match[1].isalpha():
                            month_str, day_str, year_str = match[1], match[2], match[3]
                            if int(year_str) == 2025:
                                return f"{month_str} {day_str}, {year_str}"
                        
                        # Pattern: "August 06, 2025"
                        elif len(match) == 3 and match[0].isalpha():
                            month_str, day_str, year_str = match[0], match[1], match[2]
                            if int(year_str) == 2025:
                                return f"{month_str} {day_str}, {year_str}"
                        
                        # Pattern: "08/06/2025"
                        elif len(match) == 3 and match[0].isdigit():
                            month_num, day_num, year_str = match[0], match[1], match[2]
                            if int(year_str) == 2025:
                                try:
                                    date_obj = datetime(int(year_str), int(month_num), int(day_num))
                                    return date_obj.strftime("%B %d, %Y")
                                except ValueError:
                                    continue
                    except (ValueError, IndexError):
                        continue
        
        # Method 3: Look for structured date elements
        date_selectors = [
            '.byline time',
            '.article-date',
            '.published-date', 
            '.post-date',
            'time[datetime]',
            '.entry-date',
            '.date-published'
        ]
        
        for selector in date_selectors:
            date_elem = soup.select_one(selector)
            if date_elem:
                # Try datetime attribute
                datetime_attr = date_elem.get('datetime')
                if datetime_attr:
                    try:
                        parsed_date = datetime.fromisoformat(datetime_attr.replace('Z', '+00:00'))
                        return parsed_date.strftime("%B %d, %Y")
                    except:
                        pass
                
                # Try text content
                date_text = date_elem.get_text(strip=True)
                if date_text:
                    # Look for date patterns in text
                    for pattern in byline_patterns:
                        match = re.search(pattern, date_text, re.IGNORECASE)
                        if match:
                            groups = match.groups()
                            if len(groups) >= 3:
                                try:
                                    if groups[0].isalpha():  # Month name first
                                        month_str, day_str, year_str = groups[0], groups[1], groups[2]
                                        if int(year_str) == 2025:
                                            return f"{month_str} {day_str}, {year_str}"
                                except (ValueError, IndexError):
                                    continue
        
        return None
        
    except Exception as e:
        print(f"Error extracting date from {url}: {e}")
        return None

def extract_inquirer_date_flexible(article_element, url):
    """Extract date from Inquirer article with strict validation for today/yesterday only"""
    try:
        # Current date for comparison
        today = datetime.now()
        yesterday = today - timedelta(days=1)
        
        # Method 1: Try to get actual date from the article URL (most reliable)
        print(f"🔍 Checking actual date for: {url}")
        actual_date = extract_actual_article_date(url)
        if actual_date:
            print(f"📅 Found actual article date: {actual_date}")
            # Parse the date and check if it's today or yesterday
            try:
                article_date = datetime.strptime(actual_date, "%B %d, %Y").date()
                if article_date in [today.date(), yesterday.date()]:
                    print(f"✅ Article date {actual_date} matches target dates")
                    return actual_date
                else:
                    print(f"❌ Article date {actual_date} is not from target dates (today: {today.strftime('%B %d, %Y')}, yesterday: {yesterday.strftime('%B %d, %Y')})")
                    return None  # Return None for old articles instead of current date
            except ValueError:
                print(f"⚠️ Could not parse extracted date: {actual_date}")
        
        # Method 2: Try to extract from URL pattern if available
        url_date_match = re.search(r'/(\d{4})/(\d{2})/(\d{2})/', url)
        if url_date_match:
            year, month, day = url_date_match.groups()
            # Only accept today or yesterday in 2025
            if year == '2025' and 1 <= int(day) <= 31:
                try:
                    date_obj = datetime(int(year), int(month), int(day))
                    # Check if it's today or yesterday
                    if date_obj.date() in [today.date(), yesterday.date()]:
                        print(f"✅ URL date {date_obj.strftime('%B %d, %Y')} matches target dates")
                        return date_obj.strftime("%B %d, %Y")
                    else:
                        print(f"❌ URL date {date_obj.strftime('%B %d, %Y')} is not from target dates")
                        return None
                except ValueError:
                    pass
        
        # Method 3: Look for date text in the article element - strict checking for today/yesterday
        if article_element:
            full_text = article_element.get_text()
            
            # Clean the text
            clean_text = re.sub(r'\s+', ' ', full_text).strip()
            
            # Target dates: today and yesterday
            target_dates = [
                today.strftime("%B %d, %Y"),
                yesterday.strftime("%B %d, %Y")
            ]
            
            # Look for exact target date matches
            for target_date in target_dates:
                if target_date in clean_text:
                    print(f"✅ Found target date in text: {target_date}")
                    return target_date
            
            # Try flexible patterns but only for today/yesterday
            today_patterns = [
                rf'{today.strftime("%B")}\s+{today.day},?\s*2025',
                rf'{yesterday.strftime("%B")}\s+{yesterday.day},?\s*2025',
                rf'{today.day}\s+{today.strftime("%B")}\s+2025',
                rf'{yesterday.day}\s+{yesterday.strftime("%B")}\s+2025',
            ]
            
            for pattern in today_patterns:
                match = re.search(pattern, clean_text, re.IGNORECASE)
                if match:
                    # Determine which date it matches
                    if str(today.day) in match.group() and today.strftime("%B").lower() in match.group().lower():
                        print(f"✅ Found today's date pattern: {today.strftime('%B %d, %Y')}")
                        return today.strftime("%B %d, %Y")
                    elif str(yesterday.day) in match.group() and yesterday.strftime("%B").lower() in match.group().lower():
                        print(f"✅ Found yesterday's date pattern: {yesterday.strftime('%B %d, %Y')}")
                        return yesterday.strftime("%B %d, %Y")
        
        # If no valid date found, return None instead of current date
        print(f"❌ No valid target date found for article, will be filtered out")
        return None
        
    except Exception as e:
        print(f"Error extracting Inquirer date: {e}")
        return None


def is_article_from_target_dates(published_date):
    """Check if article is from today or yesterday only (strict date filtering)"""
    if not published_date:
        print(f"📅 Skipping article - no date provided")
        return False
    
    try:
        # Dynamic target dates: today and yesterday only
        today = datetime.now()
        yesterday = today - timedelta(days=1)
        
        # Format target dates
        target_dates = [
            today.strftime("%B %d, %Y"),
            yesterday.strftime("%B %d, %Y")
        ]
        
        print(f"🎯 Target dates: {target_dates}")
        print(f"📅 Checking article date: '{published_date}'")
        
        # Clean the date string
        clean_date = published_date.strip()
        
        # Check direct match
        if clean_date in target_dates:
            print(f"✅ Including article from {clean_date}")
            return True
        
        # Try to parse and compare dates more strictly
        try:
            article_date = datetime.strptime(clean_date, "%B %d, %Y").date()
            
            # Check if it matches today or yesterday ONLY
            if article_date in [today.date(), yesterday.date()]:
                print(f"✅ Including article from {clean_date}")
                return True
            else:
                print(f"📅 Skipping article from {clean_date} (not {today.strftime('%b %d')} or {yesterday.strftime('%b %d')}, {today.year})")
                return False
                
        except ValueError:
            # Try alternative parsing
            date_match = re.search(r'(\w+)\s+(\d{1,2}),?\s+(\d{4})', clean_date)
            if date_match:
                month_str, day_str, year_str = date_match.groups()
                
                # Only accept current year
                if int(year_str) != 2025:
                    print(f"📅 Skipping article from {year_str} (not current year)")
                    return False
                
                try:
                    article_date = datetime.strptime(f"{month_str} {day_str}, {year_str}", "%B %d, %Y").date()
                    
                    # Check if it matches today or yesterday ONLY
                    if article_date in [today.date(), yesterday.date()]:
                        print(f"✅ Including article from {clean_date}")
                        return True
                    else:
                        print(f"📅 Skipping article from {clean_date} (not {today.strftime('%b %d')} or {yesterday.strftime('%b %d')}, {today.year})")
                        return False
                except ValueError:
                    pass
            else:
                print(f"⚠️ Could not parse date format: {published_date}")
                return False
        
        print(f"⚠️ Could not parse date: {published_date}")
        return False
        
    except Exception as e:
        print(f"Error checking date {published_date}: {e}")
        return False
            
    except Exception as e:
        print(f"Error checking date {published_date}: {e}")
        return False
def is_article_from_target_dates(published_date):
    """Check if article is from today or yesterday (dynamic date filtering) with enhanced validation"""
    if not published_date:
        print(f"📅 Skipping article - no valid date found")
        return False
    
    try:
        # Dynamic target dates: today and yesterday
        today = datetime.now()
        yesterday = today - timedelta(days=1)
        target_dates = [today.date(), yesterday.date()]
        
        print(f"🎯 Target dates: {[d.strftime('%B %d, %Y') for d in target_dates]}")
        print(f"📅 Checking article date: '{published_date}'")
        
        # Clean the date string first
        clean_date = published_date.strip()
        
        # Remove any extra text like "@inquirerdotnet" or "Philippine Daily Inquirer"
        clean_date = re.sub(r'@\w+|Philippine Daily Inquirer|/', '', clean_date).strip()
        
        # Try to parse the article date string
        # Handle various date formats
        date_formats = [
            "%B %d, %Y",     # August 12, 2025
            "%b %d, %Y",     # Aug 12, 2025  
            "%Y-%m-%d",      # 2025-08-12
            "%m/%d/%Y",      # 08/12/2025
            "%d %B %Y",      # 12 August 2025
        ]
        
        parsed_date = None
        
        for date_format in date_formats:
            try:
                parsed_date = datetime.strptime(clean_date, date_format).date()
                break
            except ValueError:
                continue
        
        # If standard parsing failed, try regex extraction
        if not parsed_date:
            # Look for month day, year pattern with strict validation
            date_match = re.search(r'(August|AUGUST)\s+(\d{1,2}),?\s*(2025)', clean_date, re.IGNORECASE)
            if date_match:
                month_str, day_str, year_str = date_match.groups()
                
                # Only accept 2025 and validate day
                if int(year_str) == 2025 and 1 <= int(day_str) <= 31:
                    try:
                        parsed_date = datetime.strptime(f"August {day_str}, {year_str}", "%B %d, %Y").date()
                    except ValueError:
                        print(f"📅 Invalid date format: August {day_str}, {year_str}")
                        return False
                else:
                    print(f"📅 Skipping invalid or old article from {year_str}")
                    return False
            else:
                print(f"📅 Could not parse date pattern in: '{clean_date}'")
                return False
        
        if parsed_date:
            # Additional check: reject articles more than 7 days old
            days_diff = (today.date() - parsed_date).days
            if days_diff > 7:
                print(f"📅 Skipping old article from {clean_date} (more than 7 days old)")
                return False
            
            # Check if article is from target dates (today or yesterday)
            if parsed_date in target_dates:
                print(f"✅ Including Inquirer article from {clean_date}")
                return True
            else:
                print(f"📅 Skipping Inquirer article from {clean_date} (not {today.strftime('%b %d')} or {yesterday.strftime('%b %d')}, {today.year})")
                return False
        else:
            print(f"⚠️ Could not parse Inquirer date: {published_date}")
            return False
            
    except Exception as e:
        print(f"Error checking Inquirer date {published_date}: {e}")
        return False

def scrape_inquirer_news():
    # Multiple URLs to scrape from Inquirer Business
    urls = [
        "https://business.inquirer.net/",
        "https://business.inquirer.net/category/latest-stories",
        "https://business.inquirer.net/property",
        "https://business.inquirer.net/category/latest-stories/industries",
        "https://business.inquirer.net/category/latest-stories/consumer-retail",
        "https://business.inquirer.net/category/latest-stories/tourism-and-transportation",
        "https://business.inquirer.net/category/latest-stories/economy",
        "https://business.inquirer.net/category/latest-stories/communications",
        "https://business.inquirer.net/category/latest-stories/movements"
    ]
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept-Encoding": "gzip, deflate",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    }
    
    session = requests.Session()
    news_list = []
    
    print(f"🔍 Starting Inquirer Business News Scraping...")
    
    for url_index, url in enumerate(urls, 1):
        try:
            print(f"  [{url_index}/{len(urls)}] Scraping: {url}")
            response = session.get(url, headers=headers)
            response.raise_for_status()
            
            # Add a small delay to let any JavaScript load
            time.sleep(2)
            
            soup = BeautifulSoup(response.text, "html.parser")
            
            page_news = extract_inquirer_articles(soup)
            
            # Add unique articles only
            for item in page_news:
                if not any(existing['title'] == item['title'] for existing in news_list):
                    news_list.append(item)
            
            print(f"    ✅ Found {len(page_news)} articles from this page")
            
        except Exception as e:
            print(f"    ❌ Error scraping {url}: {e}")
            continue
    
    return news_list

def extract_inquirer_articles(soup):
    """Extract articles from Inquirer page soup"""
    news_list = []
    
    # Try multiple selectors to find news articles
    selectors_to_try = [
        # Common news article selectors
        'a[href*="/business/"]',  # Links containing /business/
        'h2 a', 'h3 a', 'h4 a',  # Headlines with links
        '.entry-title a',         # Common WordPress class
        '.post-title a',          # Another common class
        'article a',              # Article links
        '.headline a',            # Headline class
        '.title a'                # Title class
    ]
    
    for selector in selectors_to_try:
        links = soup.select(selector)
        print(f"Trying selector '{selector}': found {len(links)} links")
        
        if links:
            for link in links:
                href = link.get('href', '')
                title = link.get_text(strip=True)
                
                # Filter for actual news articles
                if (title and len(title) > 10 and 
                    ('business' in href.lower() or 'inquirer.net' in href.lower() or href.startswith('/')) and
                    not any(skip in href.lower() for skip in ['javascript:', 'mailto:', '#', 'facebook', 'twitter', 'instagram'])):
                    
                    # Make relative URLs absolute
                    if href.startswith('/'):
                        href = 'https://business.inquirer.net' + href
                    elif not href.startswith('http'):
                        href = 'https://business.inquirer.net/' + href
                    
                    # Try to find category and description
                    category = None
                    description = extract_description(link)
                    
                    # Enhanced category detection
                    parent = link.find_parent()
                    if parent:
                        # Look for category elements
                        category_selectors = [
                            '[class*="category"]', '[class*="tag"]', '[class*="section"]',
                            '.cat', '.section-name', '.topic'
                        ]
                        for sel in category_selectors:
                            category_elem = parent.find(sel) or parent.select_one(sel)
                            if category_elem:
                                category = category_elem.get_text(strip=True)
                                break
                    
                    # Use intelligent categorization if no category found
                    if not category or category.lower() in ['business', 'news', '']:
                        category = categorize_news(title, description)
                    
                    # Extract actual publication date - use more flexible date extraction
                    published_date = extract_inquirer_date_flexible(link.find_parent(), href)
                    
                    # Skip articles where we couldn't extract a valid target date
                    if published_date is None:
                        print(f"    📅 Skipping Inquirer article - not from target dates: {title[:50]}...")
                        continue
                    
                    # Filter by date - strict target date filtering (today and yesterday only)
                    if not is_article_from_target_dates(published_date):
                        print(f"    📅 Skipping Inquirer article from {published_date}: {title[:50]}...")
                        continue
                    
                    print(f"    ✅ Including Inquirer article from {published_date}: {title[:50]}...")
                    
                    # Perform sentiment analysis
                    sentiment_data = get_sentiment_analysis(title + " " + (description or ""))
                    
                    news_item = {
                        "title": title,
                        "category": category or "Business",
                        "description": description or "No description available",
                        "link": href,
                        "published_date": published_date,
                        "sentiment_score": sentiment_data['sentiment_score'],
                        "sentiment_label": sentiment_data['sentiment_label'],
                        "emotion": sentiment_data['emotion'],
                        "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }
                    
                    # Avoid duplicates
                    if not any(item['title'] == title for item in news_list):
                        news_list.append(news_item)
            
            if news_list:
                break  # Stop trying other selectors if we found news
    
    return news_list

def upload_to_azure_blob(file_path, blob_name):
    """Upload file to Azure Blob Storage"""
    try:
        # Get Azure connection details from environment
        connection_string = os.getenv('AZURE_CONNECTION_STRING')
        container_name = os.getenv('AZURE_CONTAINER_NAME')
        blob_subfolder = "Data/NSI/data/Azure Databricks/Automation Scripts/News/"
        
        if not connection_string or not container_name:
            print("❌ Error: AZURE_CONNECTION_STRING or AZURE_CONTAINER_NAME not found in .env file")
            return False
        
        # Create the BlobServiceClient
        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        
        # Create the full blob path
        blob_path = blob_subfolder + blob_name
        
        # Upload the file
        with open(file_path, "rb") as data:
            blob_client = blob_service_client.get_blob_client(
                container=container_name, 
                blob=blob_path
            )
            blob_client.upload_blob(data, overwrite=True)
        
        print(f"✅ Successfully uploaded {blob_name} to Azure Blob Storage")
        print(f"📁 Container: {container_name}")
        print(f"🗂️ Path: {blob_path}")
        return True
        
    except Exception as e:
        print(f"❌ Error uploading to Azure Blob Storage: {e}")
        return False

def post_to_teams(news_items):
    """Post news summary to Microsoft Teams"""
    try:
        webhook_url = os.getenv('TEAMS_WEBHOOK_URL')
        if not webhook_url:
            print("ℹ️ Teams webhook URL not configured - skipping Teams posting")
            return False
        
        # Get top 5 news by sentiment and category diversity
        df = pd.DataFrame(news_items)
        
        # Select diverse news items (mix of categories and sentiments)
        top_news = []
        categories_used = set()
        
        # Prioritize positive and concerning news
        for sentiment in ['Positive', 'Negative', 'Neutral']:
            for _, item in df[df['sentiment_label'] == sentiment].iterrows():
                if len(top_news) >= 5:
                    break
                if item['category'] not in categories_used or len(top_news) < 3:
                    top_news.append(item)
                    categories_used.add(item['category'])
        
        # Create Teams message
        total_articles = len(news_items)
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        # Create summary statistics
        sentiment_stats = df['sentiment_label'].value_counts()
        category_stats = df['category'].value_counts()
        
        message = {
            "@type": "MessageCard",
            "@context": "http://schema.org/extensions",
            "themeColor": "0076D7",
            "summary": f"📰 Inquirer Business News Update - {total_articles} articles scraped",
            "sections": [
                {
                    "activityTitle": "📰 **Inquirer Business News Update**",
                    "activitySubtitle": f"🕒 {timestamp} | 📊 {total_articles} articles scraped",
                    "facts": [
                        {"name": "📈 Positive News", "value": str(sentiment_stats.get('Positive', 0))},
                        {"name": "📉 Negative News", "value": str(sentiment_stats.get('Negative', 0))},
                        {"name": "⚖️ Neutral News", "value": str(sentiment_stats.get('Neutral', 0))},
                        {"name": "🏷️ Categories", "value": str(len(category_stats))}
                    ],
                    "text": "**📋 Top News Headlines:**\n\n" + "\n\n".join([
                        f"**{get_emoji_for_category(item['category'])} {item['title']}**  \n"
                        f"🏷️ {item['category']} | {get_sentiment_emoji(item['sentiment_label'])} {item['sentiment_label']}  \n"
                        f"🔗 [Read More]({item['link']})"
                        for item in top_news[:5]
                    ])
                }
            ]
        }
        
        # Send to Teams
        response = requests.post(webhook_url, json=message)
        if response.status_code == 200:
            print(f"✅ Successfully posted news summary to Teams")
            return True
        else:
            print(f"❌ Failed to post to Teams: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"❌ Error posting to Teams: {e}")
        return False

def get_emoji_for_category(category):
    """Get emoji for news category"""
    emoji_map = {
        'Stock Market': '📈',
        'Banking & Finance': '🏦',
        'Energy & Utilities': '⚡',
        'Government & Policy': '🏛️',
        'Technology': '💻',
        'Real Estate': '🏠',
        'Infrastructure': '🏗️',
        'Retail & Consumer': '🛒',
        'Manufacturing': '🏭',
        'Agriculture': '🌾',
        'International Trade': '🌍',
        'Economic Indicators': '📊'
    }
    return emoji_map.get(category, '📰')

def get_sentiment_emoji(sentiment):
    """Get emoji for sentiment"""
    emoji_map = {
        'Positive': '😊',
        'Negative': '😟',
        'Neutral': '😐'
    }
    return emoji_map.get(sentiment, '😐')

if __name__ == "__main__":
    print("🔍 Starting Inquirer Business News Scraping...")
    
    news = scrape_inquirer_news()
    if news:
        # Create DataFrame and sort by date (newest first)
        df = pd.DataFrame(news)
        
        # Convert published_date to datetime for proper sorting
        try:
            df['published_date_dt'] = pd.to_datetime(df['published_date'], format='%B %d, %Y')
            # Sort by date descending (newest first)
            df = df.sort_values('published_date_dt', ascending=False)
            # Drop the helper column
            df = df.drop('published_date_dt', axis=1)
            print(f"✅ Articles sorted by date (newest first)")
        except Exception as e:
            print(f"⚠️ Could not sort by date: {e}")
        
        filename = "inquirer_news.xlsx"

        # Save locally first (Excel with table for Power Automate)
        df.to_excel(filename, index=False)
        # Add Excel table for Power Automate compatibility
        try:
            from openpyxl import load_workbook
            from openpyxl.worksheet.table import Table, TableStyleInfo
            wb = load_workbook(filename)
            ws = wb.active
            nrows = ws.max_row
            ncols = ws.max_column
            col_letters = ws.cell(row=1, column=ncols).column_letter
            table_ref = f"A1:{col_letters}{nrows}"
            table = Table(displayName="NewsTable", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
            wb.save(filename)
            print(f"📊 Saved {len(news)} news items to {filename} (with table 'NewsTable')")
        except Exception as e:
            print(f"⚠️ Could not add Excel table: {e}")
            print(f"📊 Saved {len(news)} news items to {filename}")

        # Print summary statistics
        print(f"\n📈 Summary Statistics:")
        print(f"   Total articles: {len(df)}")
        print(f"   Categories: {df['category'].nunique()}")
        print(f"   Sentiment distribution:")
        for sentiment, count in df['sentiment_label'].value_counts().items():
            print(f"     {sentiment}: {count}")

        # Upload to Azure Blob Storage
        print(f"\n☁️ Uploading to Azure Blob Storage...")
        blob_name = "inquirer_news.xlsx"
        upload_success = upload_to_azure_blob(filename, blob_name)


        if upload_success:
            print(f"✅ Complete! File uploaded to Azure successfully.")
        else:
            print(f"⚠️ Local file saved but Azure upload failed.")

    else:
        print("❌ No news found.")
